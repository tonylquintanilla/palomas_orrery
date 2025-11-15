"""
Climate Data Fetcher - Paloma's Orrery
Preserves critical climate datasets for future reference

Data preservation is climate action.
"""

import urllib.request
import requests
import json
from datetime import datetime
import os
import shutil

# Data source URLs
MAUNA_LOA_URL = "https://gml.noaa.gov/webdata/ccgg/trends/co2/co2_mm_mlo.txt"
NASA_GISS_URL = "https://data.giss.nasa.gov/gistemp/tabledata_v4/GLB.Ts+dSST.txt"
NSIDC_V4_URL = "https://noaadata.apps.nsidc.org/NOAA/G02135/seaice_analysis/Sea_Ice_Index_Monthly_Data_by_Year_G02135_v4.0.xlsx"
# NASA Science Sea Level Indicator - publicly accessible!
NASA_SEA_LEVEL_URL = "https://science.nasa.gov/earth/explore/earth-indicators/sea-level/"
# Note: Data file must be downloaded from the indicator page
# Direct link may be: https://science.nasa.gov/system/data_files/[filename].txt

# Output files
CO2_OUTPUT_FILE = "data/co2_mauna_loa_monthly.json"
TEMP_OUTPUT_FILE = "data/temperature_giss_monthly.json"
ICE_OUTPUT_FILE = "data/arctic_ice_extent_monthly.json"
SEA_LEVEL_OUTPUT_FILE = "data/sea_level_gmsl_monthly.json"

# Ocean pH Data Sources
# BCO-DMO is the most reliable for programmatic access
BCODMO_HOT_CARBONATE_URL = "https://www.bco-dmo.org/dataset/3773"  # Landing page
# Note: BCO-DMO has direct CSV exports via their API

# Alternative: Try to find direct data files from HOT
HOT_DIRECT_DATA_URLS = [
    "https://hahana.soest.hawaii.edu/hot/products/HOT_surface_CO2.txt",
    "https://hahana.soest.hawaii.edu/hot/products/products.html",
]

PH_OUTPUT_FILE = "data/ocean_ph_hot_monthly.json"

def fetch_ocean_ph_bcodmo():
    """
    Fetch ocean pH data from BCO-DMO HOT dataset
    BCO-DMO provides structured access to HOT carbonate chemistry
    """
    print("Attempting to fetch pH data from BCO-DMO...")
    
    # BCO-DMO dataset 3773 - HOT core data
    # Try their data download API
    api_url = "https://www.bco-dmo.org/dataset/3773.csv"
    
    try:
        response = requests.get(api_url, timeout=30)
        response.raise_for_status()
        print("✓ BCO-DMO download successful")
        
        lines = response.text.split('\n')
        return parse_carbonate_data(lines, source='BCO-DMO')
        
    except requests.exceptions.RequestException as e:
        print(f"✗ BCO-DMO fetch failed: {e}")
        return None

def fetch_ocean_ph_hot_direct():
    """
    Try to fetch pH data directly from HOT program
    """
    print("Attempting to fetch pH data from HOT direct...")
    
    for url in HOT_DIRECT_DATA_URLS:
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            # Check if this looks like data (not HTML)
            if 'text/plain' in response.headers.get('Content-Type', '') or \
               not response.text.strip().startswith('<'):
                print(f"✓ HOT direct download successful from {url}")
                lines = response.text.split('\n')
                return parse_carbonate_data(lines, source='HOT')
                
        except Exception as e:
            continue
    
    print("✗ No direct HOT URL worked")
    return None

def parse_carbonate_data(lines, source='Unknown'):
    """
    Parse carbonate chemistry data (flexible parser)
    Works with BCO-DMO, HOT, or similar formats
    """
    records = []
    header_found = False
    col_map = {}
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        if not line or line.startswith('!'):
            continue
        
        # Find header
        line_lower = line.lower()
        if not header_found and ('ph' in line_lower or 'year' in line_lower):
            # Parse header
            headers = line.replace('#', '').replace('\t', ' ').split()
            headers = [h.strip() for h in headers if h.strip()]
            
            # Map columns
            for idx, h in enumerate(headers):
                h_lower = h.lower()
                if 'year' in h_lower or 'yr' in h_lower:
                    col_map['year'] = idx
                elif 'month' in h_lower or 'mon' in h_lower:
                    col_map['month'] = idx
                elif 'ph' in h_lower:
                    # Prefer pH_total or pH_t
                    if 'total' in h_lower or h_lower == 'ph_t':
                        col_map['ph_total'] = idx
                    elif 'ph' not in col_map:  # Use first pH column found
                        col_map['ph'] = idx
            
            header_found = True
            print(f"✓ Identified columns: {col_map}")
            continue
        
        if not header_found or line.startswith('#'):
            continue
        
        # Parse data
        try:
            parts = line.replace('\t', ' ').split()
            parts = [p.strip() for p in parts if p.strip()]
            
            if len(parts) < 3:
                continue
            
            # Extract values
            year = None
            month = None
            ph_value = None
            
            # Get year, month, pH using column map
            if 'year' in col_map and col_map['year'] < len(parts):
                year = int(float(parts[col_map['year']]))
            
            if 'month' in col_map and col_map['month'] < len(parts):
                month = int(float(parts[col_map['month']]))
            
            # Get pH from mapped column
            ph_col = col_map.get('ph_total') or col_map.get('ph')
            if ph_col is not None and ph_col < len(parts):
                ph_value = float(parts[ph_col])
            
            # Validate
            if year and month and ph_value:
                if 1900 <= year <= 2100 and 1 <= month <= 12 and 7.0 < ph_value < 9.0:
                    key = (year, month)
                    # For now, just store first measurement per month
                    # Could enhance to average multiple measurements
                    record = {
                        'year': year,
                        'month': month,
                        'date': f"{year}-{month:02d}",
                        'ph_total': round(ph_value, 4),
                        'source': source
                    }
                    records.append(record)
        
        except (ValueError, IndexError):
            continue
    
    # Remove duplicates, keep first per month
    seen = set()
    unique_records = []
    for r in sorted(records, key=lambda x: (x['year'], x['month'])):
        key = (r['year'], r['month'])
        if key not in seen:
            unique_records.append(r)
            seen.add(key)
    
    print(f"✓ Parsed {len(unique_records)} unique monthly records")
    return unique_records

def create_ph_metadata(records):
    """Create metadata for ocean pH dataset"""
    if not records:
        return {}
    
    first = records[0]
    latest = records[-1]
    ph_change = latest['ph_total'] - first['ph_total']
    years_span = latest['year'] - first['year']
    
    return {
        'dataset_name': 'ocean_ph_hot_monthly',
        'description': 'Ocean surface pH measurements from Hawaii Ocean Time-series',
        'source': {
            'organization': 'University of Hawaii - HOT Program',
            'station': 'Station ALOHA (22°45\'N, 158°W)',
            'url': 'https://hahana.soest.hawaii.edu/hot/',
            'data_providers': [
                'BCO-DMO: https://www.bco-dmo.org/dataset/3773',
                'Scripps CO2: https://scrippsco2.ucsd.edu/data/seawater_carbon/',
                'HOT Program: https://hahana.soest.hawaii.edu/hot/'
            ],
            'citation': 'Hawaii Ocean Time-series (HOT), University of Hawaii'
        },
        'parameters': {
            'measurement': 'pH (total scale)',
            'location': 'Surface ocean (~5m depth)',
            'frequency': 'Monthly sampling',
            'method': 'Spectrophotometric or potentiometric'
        },
        'time_range': {
            'start': first['date'],
            'end': latest['date'],
            'record_count': len(records)
        },
        'statistics': {
            'first_ph': first['ph_total'],
            'latest_ph': latest['ph_total'],
            'ph_change': round(ph_change, 4),
            'years_span': years_span,
            'annual_rate': round(ph_change / years_span, 6) if years_span > 0 else 0
        },
        'context': {
            'pre_industrial_ph': 8.2,
            'current_decline': 0.1,
            'threat': 'Ocean acidification threatens marine ecosystems',
            'impact': 'pH is logarithmic: 0.1 unit drop = 30% increase in acidity',
            'rate': 'Faster than any time in 300 million years'
        },
        'last_updated': datetime.now().isoformat(),
        'cache_file': PH_OUTPUT_FILE
    }

def fetch_ocean_ph():
    """
    Main function to fetch ocean pH data
    Tries multiple sources in order of reliability
    """
    print("Fetching ocean pH data...")
    
    # Try BCO-DMO first (most reliable for automation)
    records = fetch_ocean_ph_bcodmo()
    
    # If that fails, try direct HOT sources
    if not records:
        records = fetch_ocean_ph_hot_direct()
    
    # If still no data, provide manual instructions
    if not records:
        print()
        print("✗ Automated fetch failed")
        print()
        print("MANUAL DOWNLOAD OPTION:")
        print("1. Visit: https://www.bco-dmo.org/dataset/3773")
        print("2. Click 'Get Data' → Download CSV")
        print("3. Save as 'hot_carbonate_data.txt'")
        print("4. Run: python convert_hot_ph_to_json.py")
        print()
        return []
    
    return records

def fetch_nasa_sea_level():
    """
    Fetch NASA sea level data from science.nasa.gov indicator page
    Note: Currently requires manual download from the Earth Indicators page
    Returns list of records with date and sea level
    """
    print("Fetching NASA sea level data...")
    
    # Check if user has downloaded the file
    local_file = "nasa_earthdata_sea_level_data.txt"
    
    if os.path.exists(local_file):
        print(f"✓ Found local file: {local_file}")
        try:
            with open(local_file, 'r') as f:
                lines = f.readlines()
        except Exception as e:
            print(f"✗ Error reading local file: {e}")
            return []
    else:
        print(f"✗ File not found: {local_file}")
        print(f"  Please download from: https://science.nasa.gov/earth/explore/earth-indicators/sea-level/")
        print(f"  Save as: {local_file}")
        return []
    
    print(f"✓ Read {len(lines)} lines")
    
    records = []
    
    # Parse NASA format
    # Header lines start with HDR
    # Data format: year+fraction  GMSL(cm)  GMSL_60day_smoothed(cm)
    # Example: 1993.0109589    -0.230726    -0.445896
    
    for line in lines:
        line = line.strip()
        
        # Skip empty lines and header lines
        if not line or line.startswith('HDR'):
            continue
        
        try:
            parts = line.split()
            
            if len(parts) >= 3:
                # First column is decimal year
                year_fraction = float(parts[0])
                year = int(year_fraction)
                month = int(round((year_fraction - year) * 12)) + 1
                if month > 12:
                    month = 12
                if month < 1:
                    month = 1
                
                # Second column: GMSL in centimeters
                gmsl_cm = float(parts[1])
                
                # Third column: 60-day smoothed GMSL in cm
                gmsl_smoothed_cm = float(parts[2])
                
                records.append({
                    'year': year,
                    'month': month,
                    'year_fraction': round(year_fraction, 7),
                    'gmsl_cm': round(gmsl_cm, 3),
                    'gmsl_smoothed_cm': round(gmsl_smoothed_cm, 3)
                })
        
        except (ValueError, IndexError):
            continue  # Skip malformed lines
    
    print(f"✓ Parsed {len(records)} records")
    return records

def status_print(msg):
    """Print status message with indentation"""
    print(f"  {msg}")


def fetch_mauna_loa_co2(status_callback=None):     
    """
    Fetch Mauna Loa CO2 monthly data
    Returns list of records with year, month, and CO2 concentration
    
    IMPORTANT: Field names must match existing cache file:
    - co2_ppm (not 'average')
    - co2_deseasonalized (not 'interpolated' or 'trend')
    - days, std_dev, uncertainty (always included, even if missing data = -1, -9.99, -0.99)
    """
    def status(msg):
        if status_callback:
            status_callback(msg)
        status_print(msg)
    
    status("Fetching Mauna Loa CO₂ monthly data...")
    
    try:
        response = requests.get(MAUNA_LOA_URL, timeout=30)
        response.raise_for_status()
        
        lines = response.text.strip().split('\n')
        print(f"✓ Download successful ({len(lines)} lines)")
        
        records = []
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
                
            parts = line.split()
            if len(parts) >= 5:
                try:
                    year = int(parts[0])
                    month = int(parts[1])
                    decimal_date = float(parts[2])
                    average = float(parts[3])
                    deseasonalized = float(parts[4])
                    
                    # Skip rows where CO2 is marked as missing (-99.99)
                    if average < 0:
                        continue
                    
                    # Always include all fields (matching cached format)
                    # Use placeholder values if data is missing, just like NOAA does
                    record = {
                        'year': year,
                        'month': month,
                        'decimal_date': round(decimal_date, 4),
                        'co2_ppm': round(average, 2),  # ← Field name MUST be 'co2_ppm'
                        'co2_deseasonalized': round(deseasonalized, 2),  # ← Field name MUST be 'co2_deseasonalized'
                        # Always include these fields (even if missing = -1, -9.99, -0.99)
                        'days': int(parts[5]) if len(parts) >= 6 else -1,
                        'std_dev': round(float(parts[6]), 2) if len(parts) >= 7 else -9.99,
                        'uncertainty': round(float(parts[7]), 2) if len(parts) >= 8 else -0.99
                    }
                    
                    records.append(record)
                    
                except (ValueError, IndexError):
                    continue
        
        print(f"✓ Parsed {len(records)} records")
        return records
        
    except requests.exceptions.RequestException as e:
        print(f"✗ Error fetching CO₂ data: {e}")
        return []

def fetch_nasa_giss_temperature(status_callback=None):
    """
    Fetch NASA GISS global temperature data
    Returns: (records, metadata) tuple or (None, None) on error
    """
    def status(msg):
        if status_callback:
            status_callback(msg)
        status_print(msg)
    
    status("Downloading NASA GISS temperature data...")
    
    try:
        # NASA requires a User-Agent header now
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        req = urllib.request.Request(NASA_GISS_URL, headers=headers)
        
        with urllib.request.urlopen(req, timeout=30) as response:
            data = response.read().decode('utf-8')
        
        status("✓ Download complete")
        status("Parsing temperature data...")
        
        # Rest of the function stays the same...
        records = []
        lines = data.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            try:
                parts = line.split()
                if len(parts) >= 13 and parts[0].isdigit():
                    year = int(parts[0])
                    
                    for month_idx in range(12):
                        temp_str = parts[month_idx + 1]
                        if temp_str != '***':
                            anomaly = float(temp_str) / 100.0
                            
                            records.append({
                                'year': year,
                                'month': month_idx + 1,
                                'anomaly_c': anomaly
                            })
            except (ValueError, IndexError):
                continue
        
        status(f"✓ Parsed {len(records)} temperature records")
        
        metadata = create_temperature_metadata(records)
        return records, metadata
        
    except Exception as e:
        status(f"✗ Failed to fetch temperature data: {e}")
        return None, None

def fetch_arctic_ice():
    """
    Fetch NSIDC Arctic sea ice extent data (V4 Excel format)
    Returns list of records with year, month, and ice extent
    """
    print("Fetching NSIDC Arctic sea ice extent data (V4)...")
    
    try:
        # Check if openpyxl is available
        try:
            import openpyxl
        except ImportError:
            print("✗ openpyxl not found. Install with: pip install openpyxl")
            return []
        
        # Download Excel file
        response = requests.get(NSIDC_V4_URL, timeout=60)
        response.raise_for_status()
        print(f"✓ Download successful ({len(response.content)} bytes)")
        
        # Save to temporary file
        temp_file = "temp_ice_data.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(response.content)
        
        # Load workbook
        workbook = openpyxl.load_workbook(temp_file, data_only=True)
        
        # Use NH-Extent sheet (Northern Hemisphere extent)
        if 'NH-Extent' not in workbook.sheetnames:
            print(f"✗ 'NH-Extent' sheet not found. Available sheets: {workbook.sheetnames}")
            os.remove(temp_file)
            return []
        
        sheet = workbook['NH-Extent']
        print("✓ Parsing monthly data from 'NH-Extent' sheet...")
        
        records = []
        
        # Skip header rows (first 2 rows are headers)
        for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            try:
                year = int(row[0])  # First column is year
                
                # Columns 1-12 are monthly extent values (million km²)
                months = [
                    ('January', 1), ('February', 2), ('March', 3), ('April', 4),
                    ('May', 5), ('June', 6), ('July', 7), ('August', 8),
                    ('September', 9), ('October', 10), ('November', 11), ('December', 12)
                ]
                
                for i, (month_name, month_num) in enumerate(months, start=1):
                    if i < len(row) and row[i] is not None:
                        try:
                            extent = float(row[i])
                            
                            records.append({
                                'year': year,
                                'month': month_num,
                                'month_name': month_name,
                                'extent_million_km2': round(extent, 2)
                            })
                        except (ValueError, TypeError):
                            continue
                            
            except (ValueError, IndexError, TypeError):
                continue
        
        # Clean up temp file
        os.remove(temp_file)
        
        print(f"✓ Parsed {len(records)} records (all 12 months)")
        return records
        
    except requests.exceptions.RequestException as e:
        print(f"✗ Error fetching Arctic ice data: {e}")
        return []
    except Exception as e:
        print(f"✗ Error parsing Excel file: {e}")
        if os.path.exists("temp_ice_data.xlsx"):
            os.remove("temp_ice_data.xlsx")
        return []

def create_co2_metadata(records):
    """Create metadata for CO2 dataset"""
    return {
        'dataset_name': 'mauna_loa_co2_monthly',
        'description': 'Monthly atmospheric CO₂ concentrations from Mauna Loa Observatory, Hawaii',
        'source': {
            'organization': 'NOAA Global Monitoring Laboratory',
            'url': 'https://gml.noaa.gov/ccgg/trends/',
            'data_url': MAUNA_LOA_URL,
            'citation': 'Keeling, C.D., S.C. Piper, R.B. Bacastow, et al. Atmospheric CO₂ concentrations from Mauna Loa Observatory. NOAA Global Monitoring Laboratory.'
        },
        'cached_date': datetime.now().isoformat(),
        'record_count': len(records),
        'temporal_range': {
            'start': f"{records[0]['year']}-{records[0]['month']:02d}",
            'end': f"{records[-1]['year']}-{records[-1]['month']:02d}"
        },
        'units': 'ppm (parts per million)',
        'measurement_method': 'Non-dispersive infrared spectroscopy',
        'baseline': 'Direct atmospheric measurement',
        'license': 'Public domain - U.S. Government data',
        'threat_status': 'Mauna Loa Observatory threatened with closure (August 2025 lease expiration)',
        'preservation_priority': 'CRITICAL - 67-year continuous record at risk'
    }

def create_temperature_metadata(records):
    """Create metadata for temperature dataset"""
    return {
        'dataset_name': 'nasa_giss_temperature_anomaly',
        'description': 'Global mean surface temperature anomalies (GISTEMP v4)',
        'source': {
            'organization': 'NASA Goddard Institute for Space Studies (GISS)',
            'url': 'https://data.giss.nasa.gov/gistemp/',
            'citation': 'GISTEMP Team, 2024: GISS Surface Temperature Analysis (GISTEMP), version 4. NASA Goddard Institute for Space Studies.'
        },
        'cached_date': datetime.now().isoformat(),
        'record_count': len(records),
        'temporal_range': {
            'start': f"{records[0]['year']}-{records[0]['month']:02d}",
            'end': f"{records[-1]['year']}-{records[-1]['month']:02d}"
        },
        'units': 'degrees Celsius (°C)',
        'measurement_method': 'Combined land-surface air and sea-surface water temperature anomalies',
        'baseline': '1951-1980 average (0°C reference)',
        'baseline_note': 'Pre-industrial baseline (~1850-1900) is approximately -0.3°C relative to this baseline',
        'license': 'Public domain - U.S. Government data',
        'threat_status': 'NASA Earth Science faces 52% budget cuts; GISS institutional future uncertain (August 2025)',
        'preservation_priority': 'CRITICAL - 145-year record, James Hansen legacy dataset'
    }

def create_ice_metadata(records):
    """Create metadata for Arctic sea ice dataset"""
    return {
        'dataset_name': 'arctic_sea_ice_extent',
        'description': 'Arctic sea ice extent from satellite observations (monthly data)',
        'source': {
            'organization': 'National Snow and Ice Data Center (NSIDC)',
            'url': 'https://nsidc.org/data/seaice_index/',
            'data_url': NSIDC_V4_URL,
            'citation': 'Fetterer, F., K. Knowles, W. N. Meier, M. Savoie, A. K. Windnagel, and T. Stafford. 2025. Sea Ice Index, Version 4. Boulder, Colorado USA. NSIDC: National Snow and Ice Data Center. doi:10.7265/a98x-0f50'
        },
        'cached_date': datetime.now().isoformat(),
        'record_count': len(records),
        'temporal_range': {
            'start': f"{records[0]['year']}-{records[0]['month']:02d}",
            'end': f"{records[-1]['year']}-{records[-1]['month']:02d}"
        },
        'units': 'million km²',
        'measurement_method': 'Satellite passive microwave observations (SMMR, SSM/I, SSMIS)',
        'baseline': 'Sea ice extent from 1979-present',
        'version': 'V4.0',
        'format': 'Excel (XLSX) parsed to JSON',
        'migration_note': 'NSIDC migrated from V3 (CSV/FTP) to V4 (Excel/HTTPS) in December 2023',
        'license': 'Public domain - U.S. Government data',
        'threat_status': 'Dataset downgraded to "Basic" service level; NSIDC facing funding constraints (2025)',
        'preservation_priority': 'CRITICAL - 47-year continuous satellite record, shows dramatic climate change'
    }

def create_sea_level_metadata(records):
    """Create metadata for sea level dataset"""
    return {
        'dataset_name': 'nasa_ssh_gmsl_indicator',
        'description': 'Global Mean Sea Level from NASA-SSH Simple Gridded Sea Surface Height',
        'source': {
            'organization': 'NASA Earth Indicators / PO.DAAC',
            'url': 'https://science.nasa.gov/earth/explore/earth-indicators/sea-level/',
            'data_url': 'https://science.nasa.gov/earth/explore/earth-indicators/sea-level/',
            'citation': 'NASA-SSH. 2025. Global Mean Sea Level from Simple Gridded Sea Surface Height from Standardized Reference Missions Only Version 1. PO.DAAC, CA, USA. Dataset accessed [YYYY-MM-DD] at https://doi.org/10.5067/NSIND-GMSV1.'
        },
        'cached_date': datetime.now().isoformat(),
        'record_count': len(records),
        'temporal_range': {
            'start': f"{records[0]['year']}-{records[0]['month']:02d}",
            'end': f"{records[-1]['year']}-{records[-1]['month']:02d}"
        },
        'units': 'centimeters (cm)',
        'measurement_method': 'Satellite radar altimetry (TOPEX/Poseidon, Jason series, Sentinel-6)',
        'baseline': 'Zero mean over calendar year 1993',
        'baseline_note': 'Values shifted to have zero mean over 1993. NOT adjusted for Glacial Isostatic Adjustment (GIA).',
        'processing': '60-day smoothing available; 10-day observation windows with 7-day intervals',
        'license': 'Public domain - U.S. Government data',
        'threat_status': 'NASA Earth Science missions face budget uncertainty (2025)',
        'preservation_priority': 'CRITICAL - 32-year continuous satellite record showing accelerating rise',
        'note': 'Data manually downloaded from NASA Earth Indicators page'
    }

def save_cache(output_file, records, metadata_func):
    """
    Safely save data to JSON cache with comprehensive fail-safe protection
    Follows same pattern as orbit_data_manager.py
    """
    
    if not records:
        print(f"✗ No records to save for {output_file}")
        return False
    
    # Create metadata
    metadata = metadata_func(records)
    
    # Create full data structure
    data = {
        'metadata': metadata,
        'data': records
    }
    
    # ===== FAIL-SAFE PROTECTION =====
    
    # 1. Check if existing file exists and validate
    backup_file = f"{output_file}.backup"
    emergency_file = f"{output_file}.emergency_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    existing_size = 0
    existing_count = 0
    has_existing = os.path.exists(output_file)
    
    if has_existing:
        try:
            existing_size = os.path.getsize(output_file)
            with open(output_file, 'r') as f:
                existing_data = json.load(f)
                existing_count = len(existing_data.get('data', []))
        except:
            existing_size = 0
            existing_count = 0
    
    # 2. Write to temporary file first
    temp_file = f"{output_file}.tmp"
    
    try:
        with open(temp_file, 'w') as f:
            json.dump(data, f, indent=2)
        
        # 3. Validate new file
        new_size = os.path.getsize(temp_file)
        new_count = len(records)
        
        # 4. Safety checks
        dangerous_save = False
        warning_messages = []
        
        if has_existing and new_size > 0:
            # Check for significant shrinkage (>10%)
            size_ratio = new_size / existing_size if existing_size > 0 else 1
            if size_ratio < 0.9:
                dangerous_save = True
                warning_messages.append(
                    f"File size shrinkage detected: {existing_size} -> {new_size} bytes ({size_ratio*100:.1f}%)"
                )
            
            # Check for record loss (>5 records)
            count_diff = existing_count - new_count
            if count_diff > 5:
                dangerous_save = True
                warning_messages.append(
                    f"Record count decreased: {existing_count} -> {new_count} (lost {count_diff} records)"
                )
        
        # 5. Handle dangerous saves
        if dangerous_save:
            print(f"\n⚠️  SAVE BLOCKED for {output_file}")
            for msg in warning_messages:
                print(f"   {msg}")
            
            # Create emergency backup of existing file
            if has_existing:
                shutil.copy2(output_file, emergency_file)
                print(f"   Emergency backup created: {emergency_file}")
            
            # Remove temp file
            os.remove(temp_file)
            print(f"   Keeping existing file unchanged")
            return False
        
        # 6. Create backup of existing file
        if has_existing:
            shutil.copy2(output_file, backup_file)
        
        # 7. Move temp file to final location (atomic operation)
        shutil.move(temp_file, output_file)
        
        print(f"✓ Saved successfully")
        print(f"  File size: {new_size / 1024:.1f} KB")
        print(f"  Records: {new_count}")
        
        return True
        
    except Exception as e:
        print(f"✗ Error saving {output_file}: {e}")
        
        # Clean up temp file
        if os.path.exists(temp_file):
            os.remove(temp_file)
        
        # Restore from backup if save failed
        if has_existing and os.path.exists(backup_file):
            shutil.copy2(backup_file, output_file)
            print(f"  Restored from backup")
        
        return False

def main():
    """Main function to fetch all climate data"""
    
    print("=" * 60)
    print("  Climate Data Fetcher - Paloma's Orrery")
    print("  Data Preservation is Climate Action")
    print("=" * 60)
    print()
    
    results = []
    
    # 1. Fetch CO₂ data
    print("1. Fetching Mauna Loa CO₂ monthly data...")
    co2_records = fetch_mauna_loa_co2()
    
    if co2_records:
        success = save_cache(CO2_OUTPUT_FILE, co2_records, create_co2_metadata)
        results.append(("CO₂", len(co2_records), success))
        
        if success and co2_records:
            latest = co2_records[-1]
            first = co2_records[0]
    #        increase = latest['trend'] - first['trend']
            increase = latest['co2_ppm'] - first['co2_ppm']
            years = latest['year'] - first['year']
            
    #        print(f"\nLatest CO₂ measurement: {latest['trend']} ppm")
            print(f"Latest CO₂ measurement: {latest['co2_ppm']:.2f} ppm")
            print(f"{years}-year increase: +{increase:.2f} ppm\n")
    else:
        results.append(("CO₂", 0, False))
    
    # 2. Fetch temperature data
    print("2. Fetching NASA GISS temperature data...")
#    temp_records = fetch_nasa_giss_temperature()
    temp_records, temp_metadata = fetch_nasa_giss_temperature()
    
    if temp_records:
        success = save_cache(TEMP_OUTPUT_FILE, temp_records, create_temperature_metadata)
        results.append(("Temperature", len(temp_records), success))
        
        if success and temp_records:
            latest = temp_records[-1]
            first = temp_records[0]
    #        warming = latest['anomaly_celsius'] - first['anomaly_celsius']
            warming = latest['anomaly_c'] - first['anomaly_c']
            years = latest['year'] - first['year']
            
    #        print(f"\nLatest temperature anomaly: +{latest['anomaly_celsius']}°C")
            print(f"Latest temperature anomaly: +{latest['anomaly_c']:.2f}°C")
            print(f"{years}-year warming: +{warming:.2f}°C\n")
    else:
        results.append(("Temperature", 0, False))
    
    # 3. Fetch Arctic ice data
    print("3. Fetching NSIDC Arctic sea ice extent data...")
    ice_records = fetch_arctic_ice()
    
    if ice_records:
        success = save_cache(ICE_OUTPUT_FILE, ice_records, create_ice_metadata)
        results.append(("Arctic Ice", len(ice_records), success))
        
        if success and ice_records:
            # Find September records (minimum extent)
            sept_records = [r for r in ice_records if r['month'] == 9]
            if sept_records:
                latest = sept_records[-1]
                first = sept_records[0]
                change = latest['extent_million_km2'] - first['extent_million_km2']
                years = latest['year'] - first['year']
                
                print(f"\nLatest September ice extent: {latest['extent_million_km2']} million km²")
                print(f"{years}-year change: {change:.2f} million km² ({change/first['extent_million_km2']*100:.1f}%)\n")
    else:
        results.append(("Arctic Ice", 0, False))
    
    # 4. Fetch sea level data
    print("4. Fetching NOAA STAR global mean sea level data...")
#    sea_level_records = fetch_noaa_sea_level()
    sea_level_records = fetch_nasa_sea_level()
    
    if sea_level_records:
        success = save_cache(SEA_LEVEL_OUTPUT_FILE, sea_level_records, create_sea_level_metadata)
        results.append(("Sea Level", len(sea_level_records), success))
        
        if success and sea_level_records:
            latest = sea_level_records[-1]
            first = sea_level_records[0]
    #        rise_mm = latest['gmsl_mm'] - first['gmsl_mm']
            rise_mm = (latest['gmsl_cm'] - first['gmsl_cm']) * 10  # Convert cm to mm
            years = latest['year'] - first['year']
            
            # Convert to cm for readability
            rise_cm = rise_mm / 10
            
    #        print(f"\nLatest sea level: {latest['gmsl_mm']:.1f} mm from baseline")
            print(f"\nLatest sea level: {latest['gmsl_cm']:.2f} cm from baseline")
            print(f"{years}-year rise: +{rise_cm:.1f} cm (+{rise_mm:.1f} mm)\n")
    else:
        results.append(("Sea Level", 0, False))
    
    # 5. Fetch ocean pH data
    print("5. Fetching ocean pH data...")
    ph_records = fetch_ocean_ph()
    
    if ph_records:
        success = save_cache(PH_OUTPUT_FILE, ph_records, create_ph_metadata)
        results.append(("Ocean pH", len(ph_records), success))
        
        if success and ph_records:
            latest = ph_records[-1]
            first = ph_records[0]
            ph_change = latest['ph_total'] - first['ph_total']
            years = latest['year'] - first['year']
            
            print(f"\nLatest ocean pH: {latest['ph_total']:.4f}")
            print(f"{years}-year change: {ph_change:+.4f} pH units\n")
    else:
        results.append(("Ocean pH", 0, False))
        print("Note: Ocean pH requires manual download")

    # Summary
    print("=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    
    for dataset, count, success in results:
        status = "✓" if success else "✗"
        print(f"{status} {dataset}: {count} records")
    
    print()
    print("Data preservation complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
